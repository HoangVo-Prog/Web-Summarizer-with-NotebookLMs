import logging
import os
import random
import tempfile
import undetected_chromedriver as uc
from openpyxl import load_workbook, Workbook
from newspaper import Article

from crawler.config import (
    CHROME_DRIVER_VERSION,
    HEADLESS,
)


def setup_driver():
    """
    Initializes a Chrome WebDriver instance with undetected-chromedriver.
    """
    chrome_options = uc.ChromeOptions()

    # Use a unique temporary directory for each worker to avoid conflicts
    temp_dir = tempfile.mkdtemp()
    chrome_options.add_argument(f"--user-data-dir={temp_dir}")
    chrome_options.add_argument(f"--data-path={temp_dir}")

    # General browser settings
    chrome_options.add_argument("--window-size=1920x1080")  # Set the window size explicitly for headless
    chrome_options.add_argument("--disable-gpu")  # Disable GPU to avoid rendering issues in headless mode
    chrome_options.add_argument("--disable-infobars")  # Prevents the "Chrome is being controlled by automated test software" infobar
    chrome_options.add_argument("--disable-extensions")  # Disable extensions to speed up the process
    chrome_options.add_argument("--disable-browser-side-navigation")  # Disable some side navigations
    chrome_options.add_argument("--disable-site-isolation-trials")  # To avoid some potential rendering issues
    chrome_options.add_argument("--no-sandbox")  # Disable sandboxing (useful in headless environments)
    # Headless Mode (Avoid Detection)
    chrome_options.headless = HEADLESS

    # Prevent detection by removing automation flags and using a random User-Agent
    chrome_options.add_argument("--disable-blink-features=AutomationControlled")  # Hide automation detection

    user_agents = [
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/90.0.4430.93 Safari/537.36",
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/14.0.3 Safari/605.1.15",
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:87.0) Gecko/20100101 Firefox/87.0",
        "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.101 Safari/537.36",
        "Mozilla/5.0 (Linux; Android 10; Pixel 3 XL Build/QQ1A.200205.002; wv) AppleWebKit/537.36 (KHTML, like Gecko) Version/4.0 Chrome/91.0.4472.120 Mobile Safari/537.36",
    ]
    chrome_options.add_argument(f"user-agent={random.choice(user_agents)}")
    chrome_options.add_argument("--user-data-dir=C:/Users/{userName}/AppData/Local/Google/Chrome/User Data/Profile {#}/")
    # Configure the WebDriver version if necessary
    if CHROME_DRIVER_VERSION:
        chrome_options.add_argument(f"--webdriver-version={CHROME_DRIVER_VERSION}")

    # Initialize the driver with subprocess to avoid blocking (useful in headless environments)
    try:
        driver = uc.Chrome(
            options=chrome_options,
            use_subprocess=True,
        )
        driver.maximize_window()  # Ensures window is maximized, even in headless mode
        logging.info("WebDriver initialized successfully.")

        return driver
    except Exception as e:
        logging.error(f"Failed to initialize WebDriver: {e}")
        raise


def setup_excel(excel_file):
    """
    Initializes or loads an Excel file for data storage.
    """
    try:
        if not os.path.exists(excel_file):
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Sheet1"
            sheet.append(["index", "categories", "prompt", "response", "sources", "notebook"])
            workbook.save(excel_file)
            logging.info(f"Initialized new Excel file: {excel_file}")
        else:
            workbook = load_workbook(excel_file)
            sheet = workbook.active
            logging.info(f"Loaded existing Excel file: {excel_file}")
        return workbook, sheet
    except Exception as e:
        logging.error(f"Failed to setup Excel file '{excel_file}': {e}")
        raise


def save_debugging_info(driver, workbook, debug_html="page_source_debug.html", email=None, excel_file=None):
    """
    Saves debugging information, including the current page source and workbook.
    """
    try:
        if excel_file and not email:
            workbook.save(excel_file)
            logging.info(f"Saved all prompts and responses to {excel_file}.")
        with open(debug_html, "w", encoding="utf-8") as file:
            file.write(driver.page_source)
        logging.info(f"Saved current page source to {debug_html}.")
    except Exception as e:
        logging.error(f"Failed to save debugging info: {e}")


def get_article_with_newspaper(url, language='vi'):
    try:
        article = Article(url, language=language)
        article.download()
        article.parse()
        return article.text
    except Exception as e:
        print(f"❌ Newspaper không lấy được nội dung: {e}")
        return ""

def create_txt_file(url, output_file):
    content = get_article_with_newspaper(url)
    with open(output_file, 'w', encoding='utf-8') as f:
        f.write(content)
