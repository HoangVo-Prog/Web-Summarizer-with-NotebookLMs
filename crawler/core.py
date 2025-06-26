import logging
import os
import random
import time
import re
from openpyxl import load_workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from selenium.common.exceptions import NoSuchWindowException
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from tqdm import tqdm

from crawler.config import PROMPTS_DICT, RETRIES
from crawler.exceptions import EXCEPTION_TYPES, CrawlerException
from crawler.utils import setup_driver, setup_excel, save_debugging_info


def sanitize_text(text):
    return re.sub(ILLEGAL_CHARACTERS_RE, '', text)


def perform_login(driver, email, password):
    """
    Automates the login process using the provided email and password.
    """
    try:
        logging.info("Attempting login for email: %s", email)

        # Wait for the email input field to be visible and interactable
        email_field = WebDriverWait(driver, 30).until(
            EC.visibility_of_element_located((By.ID, "identifierId"))
        )

        email_field.send_keys(email)


        next_button = WebDriverWait(driver, 30).until(
            EC.element_to_be_clickable((By.XPATH, "//*[@id='identifierNext']"))
        )
        next_button.click()
        time.sleep(random.uniform(0.1, 0.6))  # Random delay
        time.sleep(10)

        # Wait for the password input field to be visible and interactable
        password_field = WebDriverWait(driver, 30).until(
            EC.visibility_of_element_located((By.CSS_SELECTOR, 'input[type="password"]'))
        )
        password_field.send_keys(password)
        password_field.send_keys(Keys.RETURN)
        time.sleep(random.uniform(1, 2))  # Random delay
        time.sleep(5)  # Longer wait after login to ensure successful login
        logging.info("Login successful for email: %s", email)
        return 1
    except Exception as e:
        logging.error(f"Login failed for email: {email} - {e}")
        return None


def create_button(driver):
    """
       Clicks the 'Create' button in the web application.
       """
    try:
        create_button_button = WebDriverWait(driver, 40).until(
            EC.element_to_be_clickable((By.XPATH, "//button[contains(@aria-label, 'Create new notebook')]"))
        )
        create_button_button.click()
        time.sleep(2)
    except EXCEPTION_TYPES as e:
        logging.error(f"Failed to click the create button: {e}")
        raise


def upload_file(driver, folder_path):
    for file in os.listdir(folder_path):
        file_path = os.path.join(folder_path, file)
        for attempt in range(RETRIES):
            try:
                # Locate the file input button and hover over it to reveal the file input
                file_input_button = WebDriverWait(driver, 40).until(
                    EC.presence_of_element_located((By.XPATH,
                                                    "//*[contains(@id, 'mat-mdc-dialog-')]/div/div/upload-dialog/div/div[2]/upload-main-screen/div[1]/button"))
                )
                ActionChains(driver).move_to_element(file_input_button).perform()

                # Locate the file input element
                file_input = WebDriverWait(driver, 40).until(
                    EC.presence_of_element_located((By.XPATH, "//input[@type='file']"))
                )

                # Upload the file
                logging.info(f"Uploading file: {os.path.basename(file_path)}")
                file_input.send_keys(file_path)
                time.sleep(10)

                try:
                    # Check if the upload was successful
                    tick = WebDriverWait(driver, 40).until(
                        EC.presence_of_element_located((By.XPATH,
                                                        "//*[contains(@class, 'mat-mdc-checkbox-checked')]//input[not(@aria-label='Select all sources')]"))
                    )

                    if tick:
                        logging.info(f"File uploaded successfully: {os.path.basename(file_path)}")
                        time.sleep(2)
                        break
                except EXCEPTION_TYPES:
                    continue

            except EXCEPTION_TYPES:
                logging.warning(f"Attempt {attempt + 1} failed to upload the file. Retrying...")
                time.sleep(5)
        # Place the code here
        logging.info("Access adding button")
        for _ in range(RETRIES):
            try:
                upload_button = WebDriverWait(driver, 20).until(
                    EC.element_to_be_clickable((By.XPATH,
                                                "//button[@aria-label='Add source']"))
                )
                upload_button.click()
                break
            except EXCEPTION_TYPES as e:
                logging.error(f"Failed to click the upload button: {e}\n")


    logging.info("Folder upload successfully")
    return

def send_prompt(driver, prompt):
    logging.info(f"Sending prompt: {prompt}")
    prompt_input = WebDriverWait(driver, 40).until(
        EC.presence_of_element_located((By.CSS_SELECTOR, "textarea"))
    )
    prompt_input.clear()
    prompt_input.send_keys(prompt)
    prompt_input.send_keys(Keys.RETURN)
    time.sleep(10)


def get_processed_indexes(excel_file):
    """
    Reads the 'File Name' column from the Excel file and returns a set of distinct file names.
    Returns an empty set if the file doesn't exist or if no data is found.
    """
    if not os.path.exists(excel_file):
        logging.warning(f"'{excel_file}' does not exist. Returning an empty set.")
        return set()

    try:
        workbook = load_workbook(excel_file)
        sheet = workbook.active
        headers = [cell.value for cell in sheet[1]]

        processed_indexes = []

        notebook_idx = headers.index('notebook')
        for idx, row in enumerate(sheet.iter_rows(), start=1):
            if not row[notebook_idx]:
                processed_indexes.append(idx)

        return processed_indexes

    except Exception as e:
        logging.error(f"Error reading '{excel_file}': {e}")
        return list()


def get_source_div(driver, sources, number):
    small_sources = [f"___{number.text}___"]
    source_divs = None

    # Scroll the element into view using JavaScript
    driver.execute_script("arguments[0].scrollIntoView(true);", number)

    for attempt in range(RETRIES):
        try:
            ActionChains(driver).move_to_element(number).perform()

            source_divs = WebDriverWait(driver, 40).until(
                lambda container: container.find_elements(
                    By.XPATH,
                    "//*[contains(@id, 'cdk-overlay-')]//labs-tailwind-structural-element-view-v2"
                )
            )
            if source_divs:
                break
        except EXCEPTION_TYPES:
            logging.warning(f"Attempt {attempt + 1} failed to find source divs. Retrying...\n")
            time.sleep(1)

    if not source_divs:
        raise Exception("Failed to locate the file input element after multiple attempts.")

    for source_div in source_divs:
        for attempt in range(RETRIES):
            try:
                divs = WebDriverWait(source_div, 40).until(
                    lambda _: source_div.find_elements(
                        By.XPATH,
                        './/span'
                    )
                )
                small_sources.extend([div.text.strip() for div in divs if div.text.strip()])
                break
            except EXCEPTION_TYPES:
                logging.warning(f"Attempt {attempt + 1} failed to find divs. Retrying...\n")

    sources.add(sanitize_text("\n".join(small_sources)))

    ActionChains(driver).move_to_element_with_offset(number, -10, -10).perform()
    return sources


def automate_prompts_and_save(driver, folder_name, sheet, index):
    prompt_index = 0

    prompts = list(PROMPTS_DICT.values())
    categories = list(PROMPTS_DICT.keys())

    for i, prompt in tqdm(enumerate(prompts), desc="Processing prompts",
                          total=len(prompts)):
        for _ in range(RETRIES):
            try:
                send_prompt(driver, prompt)
                break
            except EXCEPTION_TYPES:
                continue

        response = ""
        numbers_set = set()
        sources = set()

        response_container = None
        response_index = prompt_index * 2 + 2
        for _ in range(RETRIES):
            try:
                response_container = WebDriverWait(driver, 20).until(
                    lambda container: container.find_elements(
                        By.XPATH,
                        f"//chat-message[{response_index}]//labs-tailwind-structural-element-view-v2/div"
                    )
                )
                break
            except EXCEPTION_TYPES as e:
                logging.error(f"Failed to locate response container: {e}")
                response_container = None

        for element in response_container:
            try:
                dot = element.text
                if not any(char.isalpha() for char in dot):
                    response += dot + " "
                    continue
            except EXCEPTION_TYPES:
                pass

            try:
                texts = element.find_elements(By.XPATH,
                                              ".//span[contains(@class, 'ng-star-inserted') and not(parent::button[@class='citation-marker'])]")
            except EXCEPTION_TYPES:
                logging.error("Failed to find text elements.")
                continue
            for text in texts:
                try:
                    number_button = text.find_element(By.XPATH,
                                                      ".//button[@class='xap-inline-dialog citation-marker ng-star-inserted']")

                    response += f" ___{number_button.text}___"
                    if number_button.text not in numbers_set:
                        numbers_set.add(number_button.text)
                        sources = get_source_div(driver, sources, number_button)

                except EXCEPTION_TYPES:
                    try:
                        if text.text == "...":
                            raise CrawlerException
                        response += " " + text.text if text.text else ""
                    except EXCEPTION_TYPES:
                        for _ in range(RETRIES):
                            try:
                                WebDriverWait(text, 20).until(
                                    EC.element_to_be_clickable(
                                        (By.XPATH, ".//button[@class='citation-marker']"))
                                )
                                citation_button = text.find_element(By.XPATH, ".//button[@class='citation-marker']")
                                citation_button.click()
                                break
                            except EXCEPTION_TYPES:
                                continue
                        try:
                            addition_sources = WebDriverWait(element, 20).until(
                                lambda container: container.find_elements(By.XPATH,
                                                                          ".//span/span/button[@class='xap-inline-dialog citation-marker ng-star-inserted']")
                            )
                        except EXCEPTION_TYPES:
                            logging.info("Cannot find addition sources")
                        else:
                            for addition_source in addition_sources:
                                response += f" ___{addition_source.text}___"
                                if addition_source.text not in numbers_set:
                                    numbers_set.add(addition_source.text)
                                    sources = get_source_div(driver, sources, addition_source)

            response += '\n'
        sources_text = "\n".join(sources)
        logging.info(f"\nReceived response: \n{response}\n")
        logging.info(f"\nSources: \n{sources_text}\n")

        # Save the prompt, file name, and response to the Excel sheet
        sheet.append([index, categories[i], prompt, response, sources_text])
        prompt_index += 1

    logging.info(f"\nFinishing f{folder_name}\n")

    # Revert and create new notebook
    driver.back()
    create_button(driver)

    return


def login_with_undetected_driver(folders_path, excel_file, email, password, driver=None, initialize_only=False,
                                 workbook=None, sheet=None, processed_indexes=None):
    if not driver:
        driver = setup_driver()  # This will now use headless=True as per the updated setup_driver function

    try:
        if initialize_only:
            driver.get("https://notebooklm.google.com?hl=en")
            logging.info("Navigated to Notebook LM")

            workbook, sheet, processed_indexes = initialize_worker(
                driver, email, password, excel_file
            )
            if not workbook and not sheet and not processed_indexes:
                logging.info("Cannot load workbook and sheet properly by not successfully sign in")
                raise CrawlerException
            return driver, workbook, sheet, processed_indexes

        # Ensure processed_files is never None
        if processed_indexes is None:
            processed_indexes = list()  # Initialize as an empty set if None

        for idx, folder in tqdm(enumerate(os.listdir(folders_path)), desc="Processing files"):
            # folder: eg 1, 2, 3,  ..., 275
            folder_path = os.path.join(folders_path, folder)
            if idx in processed_indexes:
                logging.info(f"Skipping '{folder}' as it has already been processed.")
                continue

            upload_file(driver, folder_path)

            automate_prompts_and_save(driver, folder, sheet, idx+1)

    except CrawlerException:
        logging.info("Cannot access email\n")
        pass
    except Exception as e:
        logging.error(f"An exception occurred: {e}")
    finally:
        if not initialize_only:
            if workbook:
                save_debugging_info(driver, workbook, debug_html="page_source_debug.html", excel_file=excel_file)
                try:
                    driver.close()
                    logging.info("Browser closed successfully.")
                except Exception as e:
                    logging.error(f"Failed to close the browser: {e}")

    return driver, workbook, sheet, processed_indexes  # Ensure we return these values


def initialize_worker(driver, email, password, excel_file):
    """
    Handles the setup process for each worker.
    Includes automated login, workbook initialization, and navigation to the upload area.
    """

    # Automate the login process
    perform = perform_login(driver, email, password)
    if perform is None:
        return None, None, None

    logging.info(f"Logged in successfully with email: {email}")

    # Initialize or load the workbook for the worker
    workbook, sheet = setup_excel(excel_file)
    logging.info(f"Excel workbook '{excel_file}' initialized or loaded.")

    # Fetch already processed files
    processed_indexes = get_processed_indexes(excel_file)
    logging.info(f"Processed files loaded for worker handling '{excel_file}'.")

    # Navigate to the create button and simulate any required user action
    create_button(driver)

    # New button when init
    upload_button = WebDriverWait(driver, 40).until(
            EC.element_to_be_clickable((By.XPATH, "//button[@color='primary']"))
        )
    upload_button.click()

    # Return initialized objects
    return workbook, sheet, processed_indexes
